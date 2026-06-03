from __future__ import annotations

from html import escape
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Einsatzmatrix Sanierung.xlsx"
OUTPUT = ROOT / "public" / "mockups" / "anwendungsmatrix-poster-mockups.html"


PRODUCT_FAMILIES = {
    "NEODUR HE 60 rapid": "Hartstoffestriche",
    "NEODUR HE 65": "Hartstoffestriche",
    "NEODUR HE 65 Plus": "Hartstoffestriche",
    "NEODUR HE 40": "Hartstoffestriche",
    "NEODUR Level": "Dünn- & Sichtestriche",
    "TRU PC": "Dünn- & Sichtestriche",
    "CEMENT ALL": "Reparaturmörtel",
    "MORTAR MIX": "Reparaturmörtel",
    "ASPHALT REPAIR MIX": "Verkehr & Systeme",
    "DOT Europe CONCRETE MIX": "Verkehr & Systeme",
    "KOROCRETE Schnellbeton (Innenbereiche)": "Verkehr & Systeme",
    "Rapid Set Schnellbeton (Innen- u. Außenbereiche)+": "Verkehr & Systeme",
    "DUROP Abstreumaterial (für Kunstharzbeschichtungen)": "Kunstharz",
}


REDUCED_PRODUCTS = [
    {
        "label": "HE 60 rapid",
        "members": ["NEODUR HE 60 rapid"],
        "classification": "CT-C60-F8-A6",
        "thickness": "10-60 mm",
        "load_time": "24 h",
    },
    {
        "label": "HE 65 Plus",
        "members": ["NEODUR HE 65 Plus"],
        "classification": "CT-C70-F9-A6",
        "thickness": "15-30 mm",
        "load_time": "48 h",
    },
    {
        "label": "NEODUR Level",
        "members": ["NEODUR Level"],
        "classification": "CT-C40-F10",
        "thickness": "5-10 mm",
        "load_time": "24 h",
    },
    {
        "label": "Rapid Set",
        "members": [
            "CEMENT ALL",
            "MORTAR MIX",
            "DOT Europe CONCRETE MIX",
            "Rapid Set Schnellbeton (Innen- u. Außenbereiche)+",
        ],
        "classification": "C55/67",
        "thickness": '"0"-600 mm',
        "load_time": "1 h",
    },
    {
        "label": "KOROCRETE",
        "members": ["KOROCRETE Schnellbeton (Innenbereiche)"],
        "classification": "C35/45-C50/60",
        "thickness": "projektabh.",
        "load_time": "6 h",
    },
    {
        "label": "ARM",
        "members": ["ASPHALT REPAIR MIX"],
        "classification": "n. a.",
        "thickness": "30-600 mm",
        "load_time": "30 min",
    },
]


REDUCED_FIT = {
    "Höchstbelastbare Flächen": {
        "HE 60 rapid": 2,
        "HE 65 Plus": 2,
        "Rapid Set": 1,
    },
    "Logistikflächen u. Lagerhallen": {
        "HE 60 rapid": 2,
        "HE 65 Plus": 2,
        "NEODUR Level": 1,
        "Rapid Set": 1,
        "KOROCRETE": 1,
    },
    "Montage- und Werkstattflächen": {
        "HE 60 rapid": 2,
        "HE 65 Plus": 2,
        "NEODUR Level": 1,
        "KOROCRETE": 1,
    },
    "Fachmärkte u. Fachzentren": {
        "HE 60 rapid": 1,
        "HE 65 Plus": 1,
        "NEODUR Level": 2,
    },
    "Verkehrsflächen u. Infrastruktur": {
        "HE 65 Plus": 1,
        "Rapid Set": 2,
        "KOROCRETE": 1,
        "ARM": 2,
    },
    "Parkflächen- u. Parkhäuser": {
        "HE 60 rapid": 1,
        "HE 65 Plus": 2,
        "NEODUR Level": 1,
        "Rapid Set": 1,
    },
    "Verkaufs- u. Präsentationsflächen": {
        "NEODUR Level": 2,
    },
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def product_short(name: str) -> str:
    replacements = {
        "NEODUR HE 60 rapid": "HE 60 rapid",
        "NEODUR HE 65": "HE 65",
        "NEODUR HE 65 Plus": "HE 65 Plus",
        "NEODUR HE 40": "HE 40",
        "NEODUR Level": "Level",
        "TRU PC": "TRU PC",
        "CEMENT ALL": "CEMENT ALL",
        "MORTAR MIX": "MORTAR MIX",
        "ASPHALT REPAIR MIX": "ASPHALT",
        "DOT Europe CONCRETE MIX": "DOT",
        "KOROCRETE Schnellbeton (Innenbereiche)": "KOROCRETE",
        "Rapid Set Schnellbeton (Innen- u. Außenbereiche)+": "RS Schnellbeton",
        "DUROP Abstreumaterial (für Kunstharzbeschichtungen)": "DUROP",
    }
    return replacements.get(name, name)


def load_matrix():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=False)
    sheet = workbook["Tabelle1"]

    products = []
    for col in range(4, 17):
        name = clean(sheet.cell(3, col).value)
        products.append(
            {
                "col": col,
                "name": name,
                "short": product_short(name),
                "rapid": clean(sheet.cell(4, col).value),
                "load": clean(sheet.cell(5, col).value),
                "family": PRODUCT_FAMILIES.get(name, "Sonstige"),
            }
        )

    groups = []
    current = None
    for row in range(7, 55):
        group_label = clean(sheet.cell(row, 2).value)
        usecase_label = clean(sheet.cell(row, 3).value)
        if group_label and group_label != "▪":
            current = {"label": group_label, "rows": []}
            groups.append(current)
        elif group_label == "▪" and usecase_label and current:
            cells = [bool(sheet.cell(row, product["col"]).value) for product in products]
            current["rows"].append({"label": usecase_label, "cells": cells})

    return products, groups


def marker(value: bool, mode: str = "full") -> str:
    if value:
        return '<span class="yes">✓</span>'
    if mode == "blank":
        return '<span class="blank">·</span>'
    return '<span class="no">×</span>'


def fit_marker(score: int) -> str:
    if score >= 2:
        return '<span class="yes best">✓✓</span>'
    if score == 1:
        return '<span class="yes">✓</span>'
    return ""


def speed_class(value: str) -> str:
    return {
        "30 min": "speedUltra",
        "1 h": "speedVeryFast",
        "6 h": "speedFast",
        "24 h": "speedDay",
    }.get(value, "speedNormal")


def render_reduced_matrix(products, groups) -> str:
    html = ["<table class='reducedMatrix'>"]
    html.append("<thead><tr><th class='reducedCorner'>Produkt</th>")
    for product in REDUCED_PRODUCTS:
        html.append(
            "<th class='reducedProductHead'>"
            f"<span>{escape(product['label'])}</span>"
            "</th>"
        )
    html.append("</tr></thead><tbody>")

    tech_rows = [
        ("Klassifizierung", "classification"),
        ("Schichtdicke", "thickness"),
        ("belastbar nach", "load_time"),
    ]
    for label, key in tech_rows:
        html.append("<tr class='techRow'>")
        html.append(f"<td class='techLabel'>{escape(label)}</td>")
        for product in REDUCED_PRODUCTS:
            if key == "load_time":
                html.append(
                    f"<td><span class='speedPill {speed_class(product[key])}'>{escape(product[key])}</span></td>"
                )
            else:
                html.append(f"<td>{escape(product[key])}</td>")
        html.append("</tr>")

    html.append(f"<tr class='reducedDivider'><td colspan='{len(REDUCED_PRODUCTS) + 1}'>Anwendungen</td></tr>")

    for group in groups:
        if group["label"] not in REDUCED_FIT:
            continue
        html.append("<tr class='reducedUsecaseRow'>")
        html.append(f"<td class='reducedUsecase'>{escape(group['label'])}</td>")
        for reduced_product in REDUCED_PRODUCTS:
            score = REDUCED_FIT[group["label"]].get(reduced_product["label"], 0)
            html.append(f"<td>{fit_marker(score)}</td>")
        html.append("</tr>")

    html.append("</tbody></table>")
    return "".join(html)


def render_table(products, groups, *, class_name: str, blank_no: bool = False, families: bool = False) -> str:
    no_mode = "blank" if blank_no else "full"
    family_header = ""
    if families:
        family_header += "<tr><th class='corner familyCorner'>Anwendungsfeld</th>"
        last_family = None
        span = 0
        family_cells = []
        for product in products + [{"family": None}]:
            if product["family"] == last_family:
                span += 1
                continue
            if last_family is not None:
                family_cells.append(f"<th class='familyHead' colspan='{span}'>{escape(last_family)}</th>")
            last_family = product["family"]
            span = 1
        family_header += "".join(family_cells) + "</tr>"

    html = [f"<table class='{class_name}'>"]
    html.append("<thead>")
    html.append(family_header)
    html.append("<tr><th class='corner'>Anwendungsfeld</th>")
    for product in products:
        tags = []
        if product["rapid"]:
            tags.append(("rapid", "⚡"))
        if product["load"]:
            tags.append(("load", product["load"][0]))
        tag_html = "".join(
            f"<b class='{escape(kind)}'>{escape(tag)}</b>" for kind, tag in tags
        )
        html.append(
            "<th class='productHead'>"
            f"<span>{escape(product['short'])}</span>"
            f"<small>{tag_html}</small>"
            "</th>"
        )
    html.append("</tr></thead><tbody>")

    for group in groups:
        html.append(
            f"<tr class='groupRow'><td colspan='{len(products) + 1}'>{escape(group['label'])}</td></tr>"
        )
        for row in group["rows"]:
            empty = not any(row["cells"])
            html.append("<tr class='usecaseRow emptyRow'>" if empty else "<tr class='usecaseRow'>")
            html.append(f"<td class='usecase'>{escape(row['label'])}</td>")
            for value in row["cells"]:
                html.append(f"<td>{marker(value, no_mode)}</td>")
            html.append("</tr>")
    html.append("</tbody></table>")
    return "".join(html)


def render_split(products, groups) -> str:
    split_a = products[:6]
    split_b = products[6:]

    def subset_groups(sub_products):
        result = []
        for group in groups:
            new_group = {"label": group["label"], "rows": []}
            for row in group["rows"]:
                cells = []
                for product in sub_products:
                    product_index = [p["col"] for p in products].index(product["col"])
                    cells.append(row["cells"][product_index])
                if any(cells):
                    new_group["rows"].append({"label": row["label"], "cells": cells})
            if new_group["rows"]:
                result.append(new_group)
        return result

    return (
        "<div class='splitPoster'>"
        "<div class='splitLegend'><span><b class='yes'>✓</b> geeignet</span><span><b class='blank'>·</b> nicht vorgesehen</span><span><b class='legendBadge rapid'>⚡</b> rapid / schnell</span><span><b class='legendBadge'>H</b> Belastbarkeit hoch</span><span><b class='legendBadge'>M</b> Belastbarkeit mittel</span></div>"
        "<div><h3>Boden, Estrich, Sichtfläche</h3>"
        + render_table(split_a, subset_groups(split_a), class_name="posterTable compact", blank_no=True)
        + "</div><div><h3>Reparatur, Verkehr, Systeme</h3>"
        + render_table(split_b, subset_groups(split_b), class_name="posterTable compact", blank_no=True)
        + "</div></div>"
    )


def main() -> None:
    products, groups = load_matrix()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    option_reduced = render_reduced_matrix(products, groups)
    option_a = render_table(products, groups, class_name="posterTable excelLike", blank_no=False)
    option_b = render_table(products, groups, class_name="posterTable familyLike", blank_no=True, families=True)
    option_c = render_split(products, groups)

    html = f"""<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Anwendungsmatrix Poster-Mockups</title>
  <style>
    :root {{
      --navy: #002d59;
      --cyan: #009ee3;
      --green: #009a44;
      --red: #d68a84;
      --line: #dfe7f1;
      --soft: #f4f6f9;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--navy);
      background: #f7f9fc;
    }}
    header {{
      padding: 28px 34px 18px;
      background: #fff;
      border-bottom: 1px solid var(--line);
      position: sticky;
      top: 0;
      z-index: 5;
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 30px;
      line-height: 1.1;
      letter-spacing: 0;
    }}
    .sub {{
      margin: 0;
      color: rgba(0,45,89,.68);
      font-size: 14px;
      line-height: 1.45;
    }}
    main {{
      padding: 24px 34px 60px;
      display: grid;
      gap: 34px;
    }}
    section {{
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: auto;
      box-shadow: 0 1px 2px rgba(0,0,0,.04);
    }}
    .sectionHead {{
      position: sticky;
      left: 0;
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 18px;
      padding: 18px 20px;
      background: #fff;
      border-bottom: 1px solid var(--line);
      z-index: 4;
    }}
    h2 {{
      margin: 0 0 6px;
      font-size: 20px;
      line-height: 1.2;
    }}
    .note {{
      margin: 0;
      max-width: 760px;
      color: rgba(0,45,89,.66);
      font-size: 13px;
      line-height: 1.45;
    }}
    .badge {{
      flex: 0 0 auto;
      padding: 6px 9px;
      border-radius: 5px;
      background: var(--soft);
      font-size: 11px;
      font-weight: 800;
      letter-spacing: .04em;
      text-transform: uppercase;
      color: rgba(0,45,89,.72);
    }}
    .tableWrap {{ padding: 18px 20px 24px; min-width: max-content; }}
    .posterWrap {{
      padding: 30px;
      min-width: 1540px;
      background: #111820;
      border-top: 1px solid rgba(255,255,255,.08);
    }}
    .posterCanvas {{
      width: 1500px;
      height: 500px;
      background: #fff;
      border: 1px solid rgba(255,255,255,.72);
      box-shadow: 0 18px 46px rgba(0,0,0,.34);
      display: grid;
      grid-template-rows: 84px 1fr;
      overflow: hidden;
    }}
    .posterTop {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 28px;
      padding: 8px 22px;
      border-bottom: 1px solid var(--line);
    }}
    .posterLogo {{
      display: flex;
      align-items: center;
      min-width: 250px;
    }}
    .posterLogoFull {{
      display: block;
      width: auto;
      height: 62px;
      object-fit: contain;
    }}
    .posterTitle {{
      flex: 1 1 auto;
    }}
    .posterTitle h3 {{
      margin: 0;
      font-size: 23px;
      line-height: 1;
      letter-spacing: 0;
    }}
    .posterTitle p {{
      margin: 4px 0 0;
      color: rgba(0,45,89,.66);
      font-size: 13px;
      font-weight: 700;
    }}
    .posterBody {{
      display: grid;
      grid-template-columns: 2fr 1fr;
      min-height: 0;
    }}
    .matrixPanel {{
      min-width: 0;
      overflow: hidden;
    }}
    .imagePanel {{
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      grid-template-rows: repeat(3, 1fr);
      gap: 0;
      padding: 0;
      background: #07192a;
      border-left: 1px solid var(--line);
      min-width: 0;
    }}
    .imageTile {{
      position: relative;
      overflow: hidden;
      background: #dfe7f1;
      margin: 0;
    }}
    .imageTile img {{
      width: 100%;
      height: 100%;
      object-fit: cover;
      display: block;
    }}
    .imageTile span {{
      position: absolute;
      left: 6px;
      bottom: 6px;
      padding: 3px 5px;
      background: rgba(0,45,89,.9);
      color: #fff;
      font-size: 8px;
      font-weight: 900;
      letter-spacing: .02em;
    }}
    .legend {{
      display: flex;
      gap: 18px;
      align-items: center;
      margin: 0 0 12px;
      font-size: 12px;
      color: rgba(0,45,89,.72);
    }}
    .inlineLegend {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px 18px;
      align-items: center;
      margin: 10px 0 0;
      color: rgba(0,45,89,.72);
      font-size: 12px;
    }}
    .yes {{ color: var(--green); font-size: 15px; font-weight: 900; }}
    .no {{ color: var(--red); font-size: 14px; font-weight: 900; }}
    .blank {{ color: rgba(0,45,89,.18); font-size: 13px; }}
    .posterTable {{
      border-collapse: collapse;
      table-layout: fixed;
      width: auto;
      color: var(--navy);
      font-size: 11px;
    }}
    .reducedMatrix {{
      border-collapse: collapse;
      table-layout: fixed;
      width: 1000px;
      color: var(--navy);
      font-size: 12px;
    }}
    .reducedMatrix th,
    .reducedMatrix td {{
      border: 1px solid var(--line);
      text-align: center;
      vertical-align: middle;
    }}
    .reducedCorner {{
      width: 214px;
      padding: 9px 14px;
      background: var(--navy);
      color: #fff;
      text-align: left;
      font-size: 13px;
    }}
    .reducedProductHead {{
      width: 131px;
      padding: 7px 7px 8px;
      background: #eaf8fe;
      border-bottom: 2px solid var(--cyan);
      font-size: 14px;
      font-weight: 900;
      line-height: 1.15;
    }}
    .reducedProductHead span {{
      display: block;
      font-size: 15px;
    }}
    .techRow td {{
      height: 30px;
      padding: 5px 7px;
      background: #fbfcfe;
      font-weight: 800;
    }}
    .techRow .techLabel {{
      background: #fff;
      text-align: left;
      font-weight: 900;
      color: rgba(0,45,89,.78);
    }}
    .reducedDivider td {{
      padding: 7px 14px;
      background: var(--navy);
      color: #fff;
      text-align: left;
      font-weight: 900;
      letter-spacing: .01em;
    }}
    .reducedUsecase {{
      padding: 8px 14px;
      text-align: left;
      font-weight: 900;
      line-height: 1.2;
    }}
    .reducedUsecaseRow td {{
      height: 34px;
    }}
    .reducedUsecaseRow:nth-child(even) td {{
      background-color: #fbfcfe;
    }}
    .reducedMatrix .yes {{
      font-size: 21px;
      line-height: 1;
    }}
    .reducedMatrix .best {{
      letter-spacing: -1px;
    }}
    .speedPill {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 62px;
      height: 25px;
      padding: 0 8px;
      border-radius: 999px;
      font-weight: 900;
    }}
    .speedUltra {{
      background: #ffd84d;
      color: var(--navy);
      box-shadow: 0 0 0 3px rgba(255,216,77,.26);
    }}
    .speedVeryFast {{
      background: #ffe47a;
      color: var(--navy);
    }}
    .speedFast {{
      background: rgba(255,216,77,.48);
      color: var(--navy);
    }}
    .speedDay {{
      background: #fff;
      color: var(--navy);
      border: 2px solid rgba(255,216,77,.82);
    }}
    .speedNormal {{
      background: rgba(0,45,89,.07);
      color: rgba(0,45,89,.68);
    }}
    .posterTable th, .posterTable td {{
      border: 1px solid var(--line);
      text-align: center;
      vertical-align: middle;
    }}
    .posterTable .corner {{
      position: sticky;
      left: 0;
      z-index: 3;
      width: 280px;
      min-width: 280px;
      padding: 8px 10px;
      background: var(--navy);
      color: #fff;
      text-align: left;
      font-size: 12px;
    }}
    .posterTable .productHead {{
      position: relative;
      width: 58px;
      min-width: 58px;
      height: 158px;
      background: #fff;
      border-bottom: 2px solid var(--navy);
    }}
    .posterTable .productHead > span {{
      position: absolute;
      left: 13px;
      bottom: 18px;
      width: 150px;
      transform: rotate(-48deg);
      transform-origin: left bottom;
      text-align: left;
      white-space: nowrap;
      font-weight: 900;
      font-size: 11px;
      line-height: 1.1;
    }}
    .posterTable .productHead small {{
      position: absolute;
      top: 8px;
      left: 50%;
      transform: translateX(-50%);
      display: flex;
      gap: 3px;
      justify-content: center;
    }}
    .posterTable .productHead b {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 18px;
      height: 18px;
      border-radius: 50%;
      background: #eaf8fe;
      color: var(--cyan);
      font-size: 10px;
    }}
    .posterTable .productHead b.rapid {{
      background: rgba(0,158,227,.14);
      color: var(--cyan);
      font-size: 12px;
    }}
    .posterTable .productHead b.load {{
      background: rgba(0,45,89,.08);
      color: var(--navy);
      font-size: 10px;
    }}
    .legendBadge {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 18px;
      height: 18px;
      padding: 0 5px;
      border-radius: 999px;
      background: rgba(0,45,89,.08);
      color: var(--navy);
      font-size: 10px;
      font-weight: 900;
      line-height: 1;
    }}
    .legendBadge.rapid {{
      background: rgba(0,158,227,.14);
      color: var(--cyan);
      font-size: 12px;
    }}
    .posterTable .groupRow td {{
      position: sticky;
      left: 0;
      z-index: 2;
      padding: 7px 10px;
      background: var(--navy);
      color: #fff;
      text-align: left;
      font-weight: 900;
      font-size: 12px;
      letter-spacing: .01em;
    }}
    .posterTable .usecase {{
      position: sticky;
      left: 0;
      z-index: 1;
      width: 280px;
      min-width: 280px;
      padding: 6px 10px;
      background: #fff;
      text-align: left;
      font-weight: 700;
      line-height: 1.25;
    }}
    .posterTable .usecaseRow:nth-child(even) td {{
      background-color: #fbfcfe;
    }}
    .posterTable td {{
      width: 58px;
      min-width: 58px;
      height: 28px;
      padding: 3px;
    }}
    .posterTable .emptyRow .usecase {{
      color: rgba(0,45,89,.48);
      font-style: italic;
    }}
    .familyLike .familyHead {{
      height: 28px;
      padding: 5px 8px;
      background: #eaf8fe;
      color: var(--navy);
      font-size: 11px;
      font-weight: 900;
      border-bottom: 1px solid var(--cyan);
    }}
    .familyLike .familyCorner {{
      background: var(--navy);
    }}
    .familyLike .productHead {{
      width: 48px;
      min-width: 48px;
      height: 152px;
    }}
    .familyLike .productHead > span {{
      left: 10px;
      bottom: 16px;
      width: 145px;
      transform: rotate(-64deg);
      font-size: 11.5px;
      line-height: 1.05;
    }}
    .familyLike .productHead small {{
      top: 9px;
    }}
    .familyLike td {{
      width: 48px;
      min-width: 48px;
    }}
    .splitPoster {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 22px;
      min-width: 1320px;
      padding: 18px 20px 24px;
    }}
    .splitLegend {{
      grid-column: 1 / -1;
      display: flex;
      flex-wrap: wrap;
      gap: 18px;
      align-items: center;
      margin-bottom: -2px;
      color: rgba(0,45,89,.72);
      font-size: 12px;
    }}
    .splitPoster h3 {{
      margin: 0 0 10px;
      font-size: 16px;
    }}
    .compact {{
      width: 100%;
    }}
    .compact .corner,
    .compact .usecase {{
      width: 230px;
      min-width: 230px;
    }}
    .compact .productHead {{
      width: 56px;
      min-width: 56px;
      height: 120px;
    }}
    .compact td {{
      width: 56px;
      min-width: 56px;
      height: 24px;
    }}
    @media print {{
      header {{ position: static; }}
      main {{ padding: 12px; gap: 14px; }}
      section {{ break-inside: avoid; box-shadow: none; }}
      .sectionHead {{ position: static; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Anwendungsmatrix Sanierung</h1>
    <p class="sub">Produkte, Kernwerte und Anwendungskategorien auf einen Blick.</p>
  </header>
  <main>
    <section>
      <div class="sectionHead">
        <div>
          <h2>Poster 3:1</h2>
          <p class="note">Messeposter als Beratungswerkzeug: technische Kernwerte, Geschwindigkeit und Anwendungspassung auf einer breiten Fläche.</p>
        </div>
        <span class="badge">Poster 3 rechts</span>
      </div>
      <div class="posterWrap">
        <div class="posterCanvas">
          <div class="posterTop">
            <div class="posterLogo">
              <img class="posterLogoFull" src="/images/korodur-logo-full-colour.png" alt="KORODUR">
            </div>
            <div class="posterTitle">
              <h3>Sanierungssysteme, auf die Sie bauen können.</h3>
              <p>Produkte, Tempo und Einsatzbereiche für hochbelastete Flächen.</p>
            </div>
          </div>
          <div class="posterBody">
            <div class="matrixPanel">
              {option_reduced}
            </div>
            <div class="imagePanel">
              <figure class="imageTile">
                <img src="/images/referenzen/haidberg.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/lkw-waschstrasse.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/werkstatt-neutraubling.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/theodor-heuss-bruecke.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/autohaus-versmold.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/helipad-finnland.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/zuerich-parkhaus.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/sinusfugen.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/lkw-umfahrt-darmstadt.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/haidberg-3.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/werkstatt-neutraubling-3.jpg" alt="">
              </figure>
              <figure class="imageTile">
                <img src="/images/referenzen/theodor-heuss-bruecke-2.jpg" alt="">
              </figure>
            </div>
          </div>
        </div>
      </div>
    </section>

    <section>
      <div class="sectionHead">
        <div>
          <h2>1. Excel nah: eine große Matrix</h2>
          <p class="note">Alle Produkte bleiben sichtbar. Blitz = rapid/schnell, H/M = Belastbarkeit. Sehr gut für Plakatlogik, aber im Web nur mit Horizontal-Scroll realistisch.</p>
        </div>
        <span class="badge">Maximale Vollständigkeit</span>
      </div>
      <div class="tableWrap">
        <div class="legend"><span><b class="yes">✓</b> geeignet</span><span><b class="no">×</b> nicht vorgesehen</span><span><b class="legendBadge rapid">⚡</b> rapid / schnell</span><span><b class="legendBadge">H</b> Belastbarkeit hoch</span><span><b class="legendBadge">M</b> Belastbarkeit mittel</span></div>
        {option_a}
      </div>
    </section>

    <section>
      <div class="sectionHead">
        <div>
          <h2>2. Familienlogik: Produkte oben gruppieren</h2>
          <p class="note">Immer noch eine Matrix, aber Produktfamilien strukturieren die Spalten. Leere Zellen werden ruhiger dargestellt, damit die vielen Anwendungsfelder dominieren.</p>
        </div>
        <span class="badge">Meine Empfehlung</span>
      </div>
      <div class="tableWrap">
        <div class="legend"><span><b class="yes">✓</b> geeignet</span><span><b class="blank">·</b> nicht vorgesehen</span><span><b class="legendBadge rapid">⚡</b> rapid / schnell</span><span><b class="legendBadge">H</b> Belastbarkeit hoch</span><span><b class="legendBadge">M</b> Belastbarkeit mittel</span></div>
        {option_b}
      </div>
    </section>

    <section>
      <div class="sectionHead">
        <div>
          <h2>3. Messeplakat: zwei Produktwelten</h2>
          <p class="note">Alle Anwendungsfelder bleiben sichtbar, aber Produkte werden auf zwei nebeneinanderliegende Matrizen verteilt. Das ist ruhiger für A0/A1 und vermeidet extreme Spaltenbreite.</p>
          <p class="inlineLegend"><span><b class="yes">✓</b> geeignet</span><span><b class="blank">·</b> nicht vorgesehen</span><span><b class="legendBadge rapid">⚡</b> rapid / schnell</span><span><b class="legendBadge">H</b> Belastbarkeit hoch</span><span><b class="legendBadge">M</b> Belastbarkeit mittel</span></p>
        </div>
        <span class="badge">Plakatfreundlich</span>
      </div>
      {option_c}
    </section>
  </main>
</body>
</html>"""
    OUTPUT.write_text(html, encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
